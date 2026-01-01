
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl module not found. Please install it to read Excel files (pip install openpyxl).")
    sys.exit(1)

def read_excel(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"\n{'='*50}")
    print(f"READING: {os.path.basename(file_path)}")
    print(f"{'='*50}")

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n--- SHEET: {sheet_name} ---")
            
            # Read first 50 rows to get the gist
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                if not any(row):
                    continue
                
                # Format row for readability
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                print(row_text)
                
                row_count += 1
                if row_count > 50:
                    print("... (truncated) ...")
                    break
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    files = [
        "00_管理用_AI会計システム本体.xlsx",
        "10_実務用_ワークベンチ.xlsx"
    ]
    
    for f in files:
        full_path = os.path.join(os.getcwd(), f)
        read_excel(full_path)
